"""Carga concepto → definición desde Flashcards_Derecho_Civil.xlsx."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

PARENS = re.compile(r"\s*\([^)]*\)\s*")


@dataclass(frozen=True)
class ExcelFlashcard:
    title: str
    definition: str
    materia: str | None = None


def normalize_title(text: str) -> str:
    folded = unicodedata.normalize("NFKD", (text or "").strip().lower())
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", folded).strip()


def title_base(text: str) -> str:
    return PARENS.sub(" ", normalize_title(text)).strip()


# Homologa nombres del Excel a las categorías ya usadas en la app / apuntes.
MATERIA_ALIASES: dict[str, str] = {
    "persona": "Personas",
    "personas": "Personas",
    "responsabilidad": "REX",
    "rex": "REX",
    "responsabilidad extracontractual": "REX",
    "teoria de la ley": "Teoría de la Ley",
    "teoría de la ley": "Teoría de la Ley",
    "acto juridico": "Acto Jurídico",
    "acto jurídico": "Acto Jurídico",
    "contratos": "Contratos",
    "obligaciones": "Obligaciones",
    "bienes": "Bienes",
    "familia": "Familia",
    "sucesorio": "Sucesorio",
}


def canonical_materia(materia: str | None) -> str | None:
    if not materia:
        return None
    raw = materia.strip()
    key = normalize_title(raw)
    return MATERIA_ALIASES.get(key, raw)


class ExcelFlashcardsLoader:
    # Preferir la hoja con columna Materia para categorizar.
    PREFERRED_SHEETS = (
        "Flashcards Derecho Civil",
        "Concepto-Definición",
        "Concepto-Definicion",
    )

    def load_path(self, path: Path) -> list[ExcelFlashcard]:
        return self.load_bytes(path.read_bytes())

    def load_bytes(self, data: bytes) -> list[ExcelFlashcard]:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet_name = self._pick_sheet(wb.sheetnames)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        title_idx, def_idx, materia_idx = self._column_indexes(header)

        cards: list[ExcelFlashcard] = []
        seen: set[str] = set()
        for row in rows[1:]:
            if not row:
                continue
            title = self._cell(row, title_idx)
            definition = self._cell(row, def_idx)
            if not title or not definition or len(definition) < 10:
                continue
            key = normalize_title(title)
            if key in seen:
                continue
            seen.add(key)
            materia_raw = self._cell(row, materia_idx) if materia_idx is not None else None
            cards.append(
                ExcelFlashcard(
                    title=title,
                    definition=definition,
                    materia=canonical_materia(materia_raw),
                )
            )
        return cards

    def _pick_sheet(self, names: list[str]) -> str:
        for preferred in self.PREFERRED_SHEETS:
            for name in names:
                if name.strip().lower() == preferred.lower():
                    return name
        return names[0]

    @staticmethod
    def _column_indexes(header: list[str]) -> tuple[int, int, int | None]:
        title_idx = None
        def_idx = None
        materia_idx = None
        for i, col in enumerate(header):
            if col in {"concepto", "concept", "título", "titulo", "nombre"}:
                title_idx = i
            elif col in {"definición", "definicion", "definition"}:
                def_idx = i
            elif col in {"materia", "área", "area", "subtopic"}:
                materia_idx = i
        # Fallback: hoja Concepto-Definición sin encabezados raros
        if title_idx is None:
            title_idx = 0 if "n°" not in header and "nº" not in header else 2
        if def_idx is None:
            def_idx = 1 if title_idx == 0 else 3
        return title_idx, def_idx, materia_idx

    @staticmethod
    def _cell(row: tuple, idx: int) -> str | None:
        if idx >= len(row) or row[idx] is None:
            return None
        text = str(row[idx]).strip()
        return text or None
