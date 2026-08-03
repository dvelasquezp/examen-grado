#!/usr/bin/env python3
"""Genera checklist-revision-experto.xlsx para el revisor."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def header(ws, row: int, cols: list[str], fill: str = "1E3A5F") -> None:
    for i, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_checklist_section(ws, start_row: int, title: str, items: list[str]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    row = start_row + 1
    header(ws, row, ["Ítem", "OK", "Notas"])
    row += 1
    for item in items:
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        row += 1
    return row + 1


def main() -> None:
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "checklist-revision-experto.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    severity = DataValidation(
        type="list",
        formula1='"Bloqueante,Importante,Menor"',
        allow_blank=True,
    )
    ok_val = DataValidation(type="list", formula1='"Sí,No,Parcial"', allow_blank=True)

    # --- Instrucciones ---
    ws0 = wb.active
    ws0.title = "Instrucciones"
    lines = [
        "CHECKLIST DE REVISIÓN — EXAMEN DE GRADO ORAL (U. de Chile)",
        "",
        "Materia piloto: Derecho Civil",
        "Objetivo: validar contenido y ejercicios antes de mejoras técnicas.",
        "",
        "CÓMO ABRIR LA APP",
        "1. Instalar Docker Desktop (único requisito).",
        "2. Ejecutar Iniciar-Revision.bat (Windows) o ./iniciar-revision.sh (Mac/Linux).",
        "3. Abrir http://localhost:3000",
        "",
        "QUÉ ES BORRADOR",
        "• Flashcards y juegos: revisar textos; mecánica estable.",
        "• Simulacro oral: preguntas genéricas; feedback automático NO sustituye corrección humana.",
        "• Ejemplos cotidianos: desactivados.",
        "",
        "CÓMO USAR ESTE EXCEL",
        "• Complete las hojas Checklist, Conceptos, Hallazgos y Veredicto.",
        "• Gravedad: Bloqueante / Importante / Menor.",
        "• Lo más útil: texto corregido + fuente (Cedulario, apunte, flashcard).",
    ]
    for i, line in enumerate(lines, 1):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions["A"].width = 100

    # --- Checklist ---
    ws1 = wb.create_sheet("Checklist")
    row = 1
    sections = {
        "1. Fuentes y pipeline": [
            "Documentos correctos (Flashcards, Apuntes, Guía, Cedulario)",
            "Sin PDFs duplicados u obsoletos",
            "Fragmentos de apuntes legibles (sin cortes raros)",
            "Flashcards OCR fieles al PDF",
            "Cedulario prevalece cuando hay conflicto de definición",
        ],
        "3. Flashcards (probar ~20)": [
            "Definición al voltear es la del oral",
            "Pistas Olvidé/Difícil/Bien/Fácil ayudan",
            "Sin definiciones truncadas (…)",
            "Repaso pendiente razonable",
        ],
        "4. Emparejar (~2 rondas)": [
            "Definiciones en oración completa (mayúscula → punto)",
            "Sin restos de chunk a mitad de frase",
            "Definición corresponde al concepto",
            "Longitud legible",
        ],
        "5. Completar concepto (~10)": [
            "Oraciones completas en el párrafo",
            "Contexto del mismo tema/sección",
            "Blanco pide concepto identificable",
            "Respuesta esperada es justa",
        ],
        "6. Lógica proposicional (~5)": [
            "Contexto comprensible",
            "Pregunta refleja relación jurídica real",
            "Opción correcta defendible",
            "Explicación aclara el razonamiento",
        ],
        "7. Simulacro oral": [
            "Preguntas suenan a oral chileno",
            "Conceptos relevantes para el grado",
            "Feedback no contradice la doctrina",
            "Transcripción de audio usable (si prueba audio)",
        ],
        "8. Mapa y búsqueda": [
            "Búsqueda encuentra términos habituales",
            "Resultados llevan al contenido correcto",
            "Relaciones del mapa tienen sentido pedagógico",
        ],
    }
    for title, items in sections.items():
        row = add_checklist_section(ws1, row, title, items)
    set_widths(ws1, [55, 8, 45])
    ok_val.add(f"B3:B{row}")
    ws1.add_data_validation(ok_val)

    # --- Conceptos ---
    ws2 = wb.create_sheet("Conceptos")
    header(ws2, 1, ["Concepto", "OK (Sí/No/Parcial)", "Observación", "Gravedad", "Definición sugerida", "Fuente"])
    for r in range(2, 17):
        ws2.cell(row=r, column=1, value="")
    set_widths(ws2, [22, 14, 30, 12, 40, 20])
    severity.add("D2:D16")
    ok_val2 = DataValidation(type="list", formula1='"Sí,No,Parcial"', allow_blank=True)
    ok_val2.add("B2:B16")
    ws2.add_data_validation(severity)
    ws2.add_data_validation(ok_val2)

    # --- Emparejar ---
    ws3 = wb.create_sheet("Emparejar")
    header(ws3, 1, ["Concepto", "Definición mostrada (copiar)", "Problema", "Gravedad"])
    for r in range(2, 12):
        ws3.cell(row=r, column=1, value="")
    set_widths(ws3, [22, 50, 30, 12])
    severity.add("D2:D11")
    ws3.add_data_validation(severity)

    # --- Hallazgos ---
    ws4 = wb.create_sheet("Hallazgos")
    header(
        ws4,
        1,
        [
            "Concepto / tema",
            "Ubicación",
            "Problema",
            "Texto correcto sugerido",
            "Fuente (Cedulario/apunte/flashcard)",
            "Gravedad",
        ],
    )
    ubicacion = DataValidation(
        type="list",
        formula1='"Flashcards,Emparejar,Completar,Lógica,Oral,Ficha,Búsqueda,Mapa,Otro"',
        allow_blank=True,
    )
    for r in range(2, 52):
        ws4.cell(row=r, column=1, value="")
    set_widths(ws4, [20, 14, 28, 40, 22, 12])
    ubicacion.add("B2:B51")
    severity.add("F2:F51")
    ws4.add_data_validation(ubicacion)
    ws4.add_data_validation(severity)

    # --- Cobertura ---
    ws5 = wb.create_sheet("Cobertura")
    header(
        ws5,
        1,
        ["Área / bloque", "¿Hay conceptos?", "¿Definiciones usables?", "¿Apuntes vinculados?", "Comentario"],
    )
    areas = [
        "Acto jurídico / vicios",
        "Contrato",
        "Obligaciones",
        "Responsabilidad civil",
        "Bienes / dominio",
        "Sucesorio",
        "Familia",
        "Otro: _________",
    ]
    yesno = DataValidation(type="list", formula1='"Sí,No,Parcial"', allow_blank=True)
    for i, area in enumerate(areas, 2):
        ws5.cell(row=i, column=1, value=area)
    yesno.add("B2:E9")
    ws5.add_data_validation(yesno)
    set_widths(ws5, [28, 14, 18, 18, 35])

    # --- Veredicto ---
    ws6 = wb.create_sheet("Veredicto")
    ws6["A1"] = "¿Usarías esta plataforma para estudiar hoy?"
    ws6["A2"] = "Marque una:"
    options = [
        "Sí, como herramienta principal",
        "Sí, solo flashcards/conceptos",
        "Solo después de corregir bloqueantes",
        "No todavía",
    ]
    for i, opt in enumerate(options, 3):
        ws6.cell(row=i, column=1, value=f"☐ {opt}")
    ws6["A8"] = "Top 5 prioridades de corrección"
    header(ws6, 9, ["#", "Prioridad"])
    for i in range(1, 6):
        ws6.cell(row=9 + i, column=1, value=i)
    ws6["A16"] = "Top 3 fortalezas"
    header(ws6, 17, ["#", "Fortaleza"])
    for i in range(1, 4):
        ws6.cell(row=17 + i, column=1, value=i)
    ws6["A22"] = "Comentarios libres"
    ws6.merge_cells("A23:E30")
    set_widths(ws6, [8, 70])

    wb.save(out)
    print(f"Generado: {out}")


if __name__ == "__main__":
    main()
